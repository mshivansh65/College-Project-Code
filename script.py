import openpyxl
wb = openpyxl.load_workbook(
    "E:\\Design\\Project\\data\\Uttarkashi\\downloaded\\New folder\\Accn\\Book1.xlsx")
print(type(wb))
print(wb.active.title)
# Sheet1
sh1 = wb["Sheet1"]
print(type(sh1))
print(sh1.cell(2, 3).value)
row = sh1.max_row
col = sh1.max_column
print(row)
print(col)
k = 1
t = 0
for i in range(1, row+1):
    for j in range(1, col):
        val = sh1.cell(i, j).value
        print(t, val)
        sh1.cell(row=k, column=10, value=t)
        sh1.cell(row=k, column=11, value=val)
        k = k+1
        t = t+0.02
print(k-1)
wb.save("E:\\Design\\Project\\data\\Uttarkashi\\downloaded\\New folder\\Accn\\Book1.xlsx")


